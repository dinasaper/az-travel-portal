#!/usr/bin/env python3
"""Convert all 8 Excel files to JSON for the web app."""
import openpyxl
import json
import os
from datetime import datetime, date

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
OUT_DIR = os.path.join(os.path.dirname(__file__), "static", "data")
os.makedirs(OUT_DIR, exist_ok=True)

def serialize(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val == int(val):
            return int(val)
        return round(val, 2)
    return str(val).strip()

def extract_sheet(wb, sheet_name, header_row, data_start_row, max_col=None):
    """Extract a sheet into list of dicts."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= header_row:
        return []
    headers = [serialize(c) if c else f"col_{i}" for i, c in enumerate(rows[header_row])]
    # Remove trailing empty headers
    while headers and headers[-1].startswith("col_"):
        headers.pop()
    if max_col:
        headers = headers[:max_col]
    data = []
    for row in rows[data_start_row:]:
        record = {}
        has_data = False
        for i, h in enumerate(headers):
            val = serialize(row[i] if i < len(row) else "")
            record[h] = val
            if val:
                has_data = True
        if has_data:
            data.append(record)
    return data

def convert_all():
    results = {}

    # 1. المخزن الجديد
    print("Converting المخزن الجديد...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "المخزن_الجديد.xlsx"), data_only=True)
    # code sheet - inventory items
    code_data = extract_sheet(wb, "code", 3, 4, 7)  # headers at row 4, data from row 5
    code_data = [r for r in code_data if r.get("كود") or r.get("col_0")]
    results["inventory_items"] = code_data
    # منصرف (مبيعات)
    sales = extract_sheet(wb, "منصرف(مبيعات)", 2, 3, 12)
    results["inventory_out"] = sales
    # وارد
    incoming = extract_sheet(wb, "وارد ( مدخلات المخزن)", 2, 3, 13)
    results["inventory_in"] = incoming
    # دفتر الاصناف
    ledger = extract_sheet(wb, "دفترالاصناف", 3, 4, 14)
    results["inventory_ledger"] = ledger
    wb.close()

    # 2. صيانة 2026
    print("Converting صيانة 2026...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "صيانة_2026.xlsx"), data_only=True)
    # Code sheet - vehicle codes
    maint_codes = extract_sheet(wb, "كود", 3, 4, 12)
    results["maintenance_codes"] = maint_codes
    # Maintenance records
    maint = extract_sheet(wb, "صيانة", 3, 4, 20)
    results["maintenance"] = maint
    wb.close()

    # 3. فواتير مبيعات الصاله
    print("Converting فواتير مبيعات الصاله...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "فواتير_مبيعات_الصاله.xlsx"), data_only=True)
    sales_invoices = extract_sheet(wb, "مبيعات الاصاله 2026 ", 1, 2, 28)
    results["sales_invoices"] = sales_invoices
    wb.close()

    # 4. حسابات البنوك
    print("Converting حسابات البنوك...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "حسابات_البنوك.xlsx"), data_only=True)
    qnb_az = extract_sheet(wb, "حساب ايه زد بنك QNB", 1, 3, 7)
    results["bank_qnb_az"] = qnb_az
    qnb_asala = extract_sheet(wb, "حساب الاصالة بنك QNB", 1, 3, 7)
    results["bank_qnb_asala"] = qnb_asala
    treasury = extract_sheet(wb, "الخزنة", 2, 3, 9)
    results["bank_treasury"] = treasury
    wb.close()

    # 5. وثائق اكسا
    print("Converting وثائق اكسا...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "وثائق_اكسا.xlsx"), data_only=True)
    axa = extract_sheet(wb, "وثائق اكسا للتأمين", 0, 1, 17)
    results["insurance_axa"] = axa
    wb.close()

    # 6. وثائق اورينت
    print("Converting وثائق اورينت...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "وثائق_اورينت.xlsx"), data_only=True)
    orient1 = extract_sheet(wb, "وثائق اورينت (2)", 0, 1, 17)
    orient2 = extract_sheet(wb, "وثائق اورينت", 0, 1, 17)
    results["insurance_orient"] = orient1
    results["insurance_orient2"] = orient2
    wb.close()

    # 7. التشغيل
    print("Converting التشغيل...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "التشغيل.xlsx"), data_only=True)
    op_codes = extract_sheet(wb, "كود", 3, 4, 18)
    results["operation_codes"] = op_codes
    op_main = extract_sheet(wb, "تشغيل ", 3, 4, 20)
    results["operations"] = op_main
    wb.close()

    # 8. الخزنة
    print("Converting الخزنة...")
    wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "الخزنة.xlsx"), data_only=True)
    # العهد - main treasury
    ahd_headers = extract_sheet(wb, "العهد", 7, 8, 12)
    results["treasury_main"] = ahd_headers
    # السلف
    salaf = extract_sheet(wb, "السلف", 3, 4, 12)
    results["treasury_salaf"] = salaf
    # عهد البنزين
    benz = extract_sheet(wb, "عهد البنزين", 3, 4, 12)
    results["treasury_fuel"] = benz
    # عهدة صيانة
    maint_t = extract_sheet(wb, "عهدة صيانة", 4, 5, 18)
    results["treasury_maintenance"] = maint_t
    # دفعات تحت الحساب
    payments = extract_sheet(wb, "دفعات تحت الحساب", 2, 3, 14)
    results["treasury_payments"] = payments
    wb.close()

    # Save all
    for key, data in results.items():
        out_path = os.path.join(OUT_DIR, f"{key}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=None)
        print(f"  {key}: {len(data)} records -> {out_path}")

    # Create summary stats
    summary = {
        "files": {
            "inventory_items": len(results.get("inventory_items", [])),
            "inventory_out": len(results.get("inventory_out", [])),
            "inventory_in": len(results.get("inventory_in", [])),
            "maintenance": len(results.get("maintenance", [])),
            "sales_invoices": len(results.get("sales_invoices", [])),
            "bank_qnb_az": len(results.get("bank_qnb_az", [])),
            "bank_qnb_asala": len(results.get("bank_qnb_asala", [])),
            "bank_treasury": len(results.get("bank_treasury", [])),
            "insurance_axa": len(results.get("insurance_axa", [])),
            "insurance_orient": len(results.get("insurance_orient", [])),
            "operations": len(results.get("operations", [])),
            "treasury_main": len(results.get("treasury_main", [])),
            "treasury_salaf": len(results.get("treasury_salaf", [])),
            "treasury_fuel": len(results.get("treasury_fuel", [])),
            "treasury_maintenance": len(results.get("treasury_maintenance", [])),
            "treasury_payments": len(results.get("treasury_payments", [])),
        }
    }
    with open(os.path.join(OUT_DIR, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False)
    print("\nDone! Summary:", json.dumps(summary, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    convert_all()
